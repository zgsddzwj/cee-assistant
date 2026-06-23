# utils/exporter.py
"""数据导出器 - 支持多种格式导出

特性：
1. CSV 导出
2. Excel 导出
3. JSON 导出（格式化）
4. 批量导出
"""

import csv
import json
import os
from typing import Dict, List, Any, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed, Excel export disabled")

class DataExporter:
    """数据导出器"""
    
    def __init__(self, output_dir: str = './exports'):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export_csv(self, data: List[Dict], filename: str, fieldnames: Optional[List[str]] = None) -> str:
        """导出为 CSV 文件"""
        if not data:
            logger.warning("No data to export")
            return ""
        
        filepath = os.path.join(self.output_dir, filename)
        
        # 自动提取字段名
        if not fieldnames:
            fieldnames = list(data[0].keys())
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        
        logger.info(f"[Exporter] CSV exported: {filepath} ({len(data)} rows)")
        return filepath
    
    def export_excel(self, data: List[Dict], filename: str, sheet_name: str = "Data") -> str:
        """导出为 Excel 文件"""
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl not installed, cannot export Excel")
            return ""
        
        if not data:
            logger.warning("No data to export")
            return ""
        
        filepath = os.path.join(self.output_dir, filename)
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        fieldnames = list(data[0].keys())
        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, item in enumerate(data, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                value = item.get(field, "")
                # 处理列表和字典
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        for col_idx, field in enumerate(fieldnames, 1):
            max_length = len(field)
            for item in data:
                val_len = len(str(item.get(field, "")))
                if val_len > max_length:
                    max_length = val_len
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        wb.save(filepath)
        logger.info(f"[Exporter] Excel exported: {filepath} ({len(data)} rows)")
        return filepath
    
    def export_json(self, data: Any, filename: str, pretty: bool = True) -> str:
        """导出为格式化的 JSON 文件"""
        filepath = os.path.join(self.output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            if pretty:
                json.dump(data, f, ensure_ascii=False, indent=2)
            else:
                json.dump(data, f, ensure_ascii=False)
        
        logger.info(f"[Exporter] JSON exported: {filepath}")
        return filepath
    
    def export_scores(self, scores: List[Dict], format: str = "csv") -> str:
        """导出分数线数据"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == "csv":
            return self.export_csv(scores, f"scores_{timestamp}.csv")
        elif format == "excel" and EXCEL_AVAILABLE:
            return self.export_excel(scores, f"scores_{timestamp}.xlsx", "分数线")
        elif format == "json":
            return self.export_json(scores, f"scores_{timestamp}.json")
        else:
            logger.error(f"Unsupported format: {format}")
            return ""
    
    def export_universities(self, universities: List[Dict], format: str = "csv") -> str:
        """导出院校数据"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 展平专业数据
        flat_data = []
        for uni in universities:
            for major in uni.get("majors", []):
                flat_data.append({
                    "院校名称": uni.get("university_name", ""),
                    "院校层次": uni.get("tier", ""),
                    "省份": uni.get("province", ""),
                    "专业名称": major.get("name", ""),
                    "专业代码": major.get("code", ""),
                    "学科门类": major.get("category", ""),
                    "录取分数": major.get("admission_score", ""),
                    "招生人数": major.get("plan_count", ""),
                    "学费": major.get("tuition", ""),
                    "学制": major.get("duration", ""),
                    "热门": "是" if major.get("hot") else "否",
                })
        
        if format == "csv":
            return self.export_csv(flat_data, f"universities_{timestamp}.csv")
        elif format == "excel" and EXCEL_AVAILABLE:
            return self.export_excel(flat_data, f"universities_{timestamp}.xlsx", "院校专业")
        elif format == "json":
            return self.export_json(universities, f"universities_{timestamp}.json")
        else:
            logger.error(f"Unsupported format: {format}")
            return ""
    
    def batch_export(self, data_dict: Dict[str, List[Dict]], format: str = "csv") -> Dict[str, str]:
        """批量导出多个数据集"""
        results = {}
        for name, data in data_dict.items():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if format == "csv":
                filepath = self.export_csv(data, f"{name}_{timestamp}.csv")
            elif format == "excel" and EXCEL_AVAILABLE:
                filepath = self.export_excel(data, f"{name}_{timestamp}.xlsx", name)
            elif format == "json":
                filepath = self.export_json(data, f"{name}_{timestamp}.json")
            else:
                continue
            results[name] = filepath
        return results


# 便捷函数
def export_data(data: List[Dict], format: str = "csv", output_dir: str = "./exports") -> str:
    """便捷导出函数"""
    exporter = DataExporter(output_dir)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if format == "csv":
        return exporter.export_csv(data, f"export_{timestamp}.csv")
    elif format == "excel":
        return exporter.export_excel(data, f"export_{timestamp}.xlsx")
    elif format == "json":
        return exporter.export_json(data, f"export_{timestamp}.json")
    else:
        raise ValueError(f"Unsupported format: {format}")
